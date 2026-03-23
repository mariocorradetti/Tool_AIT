import openpyxl
from datetime import datetime

class ExcelManager:
    def __init__(self):
        self.wb = None
        self.ws = None
        self.path = None

    def load(self, path):
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb.active
        self.path = path

    def get_step(self, row):
        if not self.ws or row < 2:
            return None
        step_number = self.ws[f"A{row}"].value
        if step_number is None:
            return None
        return {
            "number": step_number,
            "description": self.ws[f"B{row}"].value,
            "expected": self.ws[f"C{row}"].value,
            "measured": self.ws[f"D{row}"].value,
            "remarks": self.ws[f"E{row}"].value,
            "signature": self.ws[f"H{row}"].value
        }

    def sign_step(self, row, username, measured, remark):
        now = datetime.now()
        if self.ws[f"D{row}"].value != measured:
            self.ws[f"D{row}"] = measured

        old_remark = self.ws[f"E{row}"].value or ""
        if remark:
            combined_remark = f"{old_remark}\n[{now} - {username}] {remark}" if old_remark else f"[{now} - {username}] {remark}"
            self.ws[f"E{row}"] = combined_remark

        self.ws[f"H{row}"] = username
        self.ws[f"G{row}"] = now
        self.wb.save(self.path)

    def find_last_signed(self):
        if not self.ws:
            return 1
        row = 2
        last_signed = 1
        while True:
            step = self.ws[f"A{row}"].value
            if step is None:
                break
            if self.ws[f"H{row}"].value:
                last_signed = row
            row += 1
        return last_signed